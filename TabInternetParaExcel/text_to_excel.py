import openpyxl
from openpyxl.utils import get_column_letter

def text_to_excel():
    try:
        # Solicita a quantidade de colunas
        num_columns = int(input("Digite a quantidade de colunas: "))
        if num_columns < 1:
            print("Erro: A quantidade de colunas deve ser pelo menos 1.")
            return

        # Lê o texto de entrada do arquivo entrada.txt
        try:
            with open('entrada.txt', 'r', encoding='utf-8') as f:
                lines = [line.strip() for line in f if line.strip()]
        except FileNotFoundError:
            print("Erro: O arquivo 'entrada.txt' não foi encontrado.")
            return
        except Exception as e:
            print(f"Erro ao ler o arquivo 'entrada.txt': {str(e)}")
            return

        if not lines:
            print("Erro: O arquivo 'entrada.txt' está vazio ou não contém texto válido.")
            return

        # Verifica se há linhas suficientes para o cabeçalho
        if len(lines) < num_columns:
            print(f"Erro: O arquivo deve ter pelo menos {num_columns} linhas para o cabeçalho.")
            return

        # Extrai o cabeçalho (primeiras num_columns linhas)
        header = lines[:num_columns]

        # Extrai os dados (grupos de num_columns linhas)
        data_rows = []
        for i in range(num_columns, len(lines), num_columns):
            row = lines[i:i + num_columns]
            if len(row) == num_columns:
                data_rows.append(row)
            else:
                print(f"Aviso: Linhas incompletas ignoradas a partir do índice {i}: {row}")

        # Cria a planilha Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Adiciona o cabeçalho
        for col_idx, value in enumerate(header, start=1):
            ws[f"{get_column_letter(col_idx)}1"] = value

        # Adiciona os dados
        for row_idx, row in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws[f"{get_column_letter(col_idx)}{row_idx}"] = value

        # Salva o arquivo Excel como saida.xlsx
        output_file = "saida.xlsx"
        wb.save(output_file)
        print(f"Arquivo Excel '{output_file}' gerado com sucesso!")
        print("\nPrévia da tabela:")
        print("\t".join(header))
        for row in data_rows:
            print("\t".join(row))

    except ValueError:
        print("Erro: A quantidade de colunas deve ser um número inteiro.")
    except Exception as e:
        print(f"Erro inesperado: {str(e)}")

if __name__ == "__main__":
    text_to_excel()